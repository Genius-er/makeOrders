import openpyxl

# 加载模板
wb = openpyxl.load_workbook('templete.xlsx')
ws = wb['Sheet1']

print('=== L列公式检查 ===')
for row in range(1, 20):
    cell = ws.cell(row=row, column=12)  # L列是第12列
    print(f'L{row}: value={repr(cell.value)}, data_type={cell.data_type}')

print('\n=== 删除第9-10行后的公式变化 ===')
ws.delete_rows(9, 2)
for row in range(1, 20):
    cell = ws.cell(row=row, column=12)
    print(f'L{row}: value={repr(cell.value)}, data_type={cell.data_type}')