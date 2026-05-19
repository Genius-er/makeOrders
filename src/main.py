import openpyxl
import os
import requests
from openpyxl.drawing.image import Image
import PIL
import urllib.request
from openpyxl.styles import Alignment
import re
from openpyxl.styles import Font
from openpyxl.cell.cell import get_column_letter
import time
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils.units import cm_to_EMU
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, OneCellAnchor, AnchorMarker
import copy
import io
import sys

def get_resource_path(relative_path):
    """
    获取资源文件的绝对路径，支持开发环境和打包后的环境
    """
    try:
        # PyInstaller 打包后的临时目录
        base_path = sys._MEIPASS
    except AttributeError:
        # 开发环境
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

def get_base_path():
    """
    获取程序运行的基础路径
    """
    if getattr(sys, 'frozen', False):
        # 打包后的环境
        return os.path.dirname(sys.executable)
    else:
        # 开发环境
        return os.path.abspath(".")

TEMPLETE_PATH = get_resource_path(r"templete.xlsx")  # 最终订单模板路径（从打包资源获取）
RESOUCE_PATH = os.path.join(get_base_path(), "店小秘导出订单表")  # 店小秘导出订单表文件夹（在exe所在目录）
OUTPUT_PATH = os.path.join(get_base_path(), "工具输出工厂单")  # 输出文件夹（在exe所在目录）


def main():
    print("start making orders!!!!!")

    start_time = time.time()

    # 确保店小秘导出订单表文件夹存在
    if not os.path.exists(RESOUCE_PATH):
        os.makedirs(RESOUCE_PATH)
        print(f"已创建文件夹: {RESOUCE_PATH}")

    # 确保输出文件夹存在
    if not os.path.exists(OUTPUT_PATH):
        os.makedirs(OUTPUT_PATH)
        print(f"已创建文件夹: {OUTPUT_PATH}")

    # 遍历 RESOUCE_PATH 文件夹下所有文件，找出xlsx文件，并打印不带后缀文件名
    for file in os.listdir(RESOUCE_PATH):
        if file.endswith(".xlsx"):
            file_name = file[:-5]
            filePath = os.path.join(RESOUCE_PATH, file)
            print("=======开始===：", filePath)

            # 读取店小秘导出订单表格
            workbookDianXiaoMi = openpyxl.load_workbook(filePath)
            worksheetDianXiaoMi = workbookDianXiaoMi["order_"]

            # 店小秘表头
            sheetHead = []
            for cell in worksheetDianXiaoMi[1]:
                sheetHead.append(cell.value)

            # 遍历店小秘中的订单
            rawData = []
            skuColor = {}
            for row in worksheetDianXiaoMi.iter_rows(min_row=2):
                if all(cell.value is None for cell in row):
                    continue # 空行跳过
                
                productSpecifications = row[sheetHead.index("产品规格")].value

                # 支持的尺码配置 - 同时支持英文格式(Size:XXX Color:XXX)和中文格式(颜色:XXX 尺寸:XXX)
                size = "UNKNOWN"
                color = "UNKNOWN"
                
                # 尝试匹配英文格式 Size:XXX
                sizeMatches = re.findall(r'Size:(?:xs|s|m|l|xl|xxl|xxxl|xxxxl|2xl|3xl|4xl|5xl)?', productSpecifications, re.IGNORECASE)
                if sizeMatches and sizeMatches[0]:
                    sizePart = sizeMatches[0]
                    size = sizePart.split(":")[1].upper() if len(sizePart.split(":")) > 1 else "UNKNOWN"
                    if size == "":
                        size = "UNKNOWN"
                    # 提取颜色（英文格式）
                    colorPart = productSpecifications.replace(sizePart, "")
                    colorMatches = re.findall(r'Color:(.+?)(?:Size|$)', colorPart, re.IGNORECASE)
                    if colorMatches:
                        color = colorMatches[0].strip().replace(" ", "_")
                    else:
                        colorParts = colorPart.split(":")
                        color = colorParts[1].strip().replace(" ", "_") if len(colorParts) > 1 else "UNKNOWN"
                else:
                    # 尝试匹配中文格式 尺寸:XXX
                    sizeMatches = re.findall(r'尺寸[:：]\s*(xs|s|m|l|xl|xxl|xxxl|xxxxl|2xl|3xl|4xl|5xl)', productSpecifications, re.IGNORECASE)
                    if sizeMatches:
                        size = sizeMatches[0].upper()
                    else:
                        # 尝试更宽松的中文尺寸匹配
                        sizeMatches = re.findall(r'尺寸[:：]\s*(\w+)', productSpecifications)
                        if sizeMatches:
                            size = sizeMatches[0].upper()
                
                    # 提取颜色（中文格式）
                    colorMatches = re.findall(r'颜色[:：]\s*([^\n\r]+)', productSpecifications)
                    if colorMatches:
                        color = colorMatches[0].strip().replace(" ", "_")
                
                # size转换为同一格式，大写，XXXL改成3XL
                num_x = len(re.findall(r'X', size))
                if num_x > 1 and size != "UNKNOWN":
                    size = size.replace('X' * num_x, str(num_x) + "X")
                
                # 如果仍然无法解析，跳过此行
                if size == "UNKNOWN" and color == "UNKNOWN":
                    print(f"警告：无法解析产品规格 '{productSpecifications}'，跳过此行")
                    continue

                goodsId = row[sheetHead.index("产品ID")].value
                rawDataItem = {}

                goodsNum = row[sheetHead.index("产品总数")].value
                if goodsId + color in skuColor:
                    rawDataItem = rawData[skuColor[goodsId+color]]
                    if size in rawDataItem["sizeNum"]:
                        rawDataItem["sizeNum"][size] = int(rawDataItem["sizeNum"][size] + goodsNum)
                    else:
                        rawDataItem["sizeNum"][size] = int(goodsNum)
                else:
                    rawDataItem = {
                        "goodsName": row[sheetHead.index("产品名称")].value,
                        "sizeNum": {size: int(goodsNum)},
                        "imgUrl": row[sheetHead.index("图片网址")].value,
                        "sku": goodsId,
                        "color": color,
                    }
                    skuColor[goodsId + color] = len(rawData)
                    rawData.append(rawDataItem)

            genfinalOrder(rawData, file_name)

    end_time = time.time()
    elapsed_time = end_time - start_time
    print("脚本总用时: {} 分 {} 秒".format(int(elapsed_time/60), int(elapsed_time%60)))
    print("输出文件在 工具输出工厂单 这个文件夹里")



def get_absolute(worksheet, row, col):
    """
    获取单元格的右下方绝对位置（单位：像素），及单元格的宽高
    """
    x = 0
    y = 0
    # get_column_letter(int)把整数转换为Excel中的列索引
    col_letter = get_column_letter(col)
    # 获取每列的列宽
    width = worksheet.column_dimensions[col_letter].width
    # 计算第一列到目标列的总宽
    for i in range(col):
        col_letter = get_column_letter(i + 1)
        fcw = worksheet.column_dimensions[col_letter].width
        x += fcw
    # 如果Excel中高为默认值时，openpyxl却没有值为NoneValue，这一点我很奇怪。
    if not worksheet.row_dimensions[col].height:
        worksheet.row_dimensions[col].height = 13.5
        height = 13.5  # Excel默认列宽为13.5
    else:
        height = worksheet.row_dimensions[col].height
    # 计算第一行到目标行的总高
    for j in range(row):
        if not worksheet.row_dimensions[j + 1].height:
            worksheet.row_dimensions[j + 1].height = 13.5
            fch = 13.5
        else:
            fch = worksheet.row_dimensions[j + 1].height
        y += fch 
    # 把高单位转换为像素
    height = (height * 18) // 13.5  # 一个单元格高为13.5，像素为18
    # 把宽单位转换为像素
    width = (width * 72) // 9  # 一个单元格为宽为9，像素为72
    x = (x * 72) // 9
    y = (y * 18) // 13.5
    return x, y, width, height

def inster_image(worksheet, start_row, start_col, hieght, image_url, image_size=None):

    img = Image(image_url)
    # 将图像转换为JPEG格式
    # img = img.convert("RGB")
    img.height, img.width = image_size
    col_letter = get_column_letter(start_col)
    width = worksheet.column_dimensions[col_letter].width
    c2e = cm_to_EMU
    p2e = pixels_to_EMU
    size = XDRPositiveSize2D(p2e(img.height), p2e(img.width))
    image_cell, mod = divmod(img.height, 20)
    image_cell = image_cell+1 if mod != 0 else image_cell
    if (hieght - image_cell) % 2 == 0:
        start = start_row + (hieght - image_cell) // 2
        rowOff = c2e((0.6 * 49.77) / 99)
    else:
        start = start_row + (hieght - image_cell) // 2 + 1
        rowOff = c2e((0.4 * 49.77) / 99)
    colOff = (c2e(width + 1) - c2e(img.width / 72 * 13)) / 9
    marker = AnchorMarker(col=start_col - 1, colOff=colOff, row=start - 1, rowOff=rowOff)
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    worksheet.add_image(img)


# 输入从店小秘导出的订单中原始数据，生成最终表格
def genfinalOrder(rawData, file_name):
    print("start gen Order-------------")

    # 读取最终订单模板
    workbookDianTemplete = openpyxl.load_workbook(TEMPLETE_PATH)
    worksheetTemplete = workbookDianTemplete["Sheet1"]

    goodsStartRowNum =11 # 商品开始的行数，每2行为一个产品（保持原有位置，避免公式错位）

    # 遍历原始数据，写入到模板表格
    # 模板表格产品表头
    goodsTableHead = []
    for cell in worksheetTemplete[7]:
        goodsTableHead.append(cell.value)
    
    for i in range(len(rawData) + 200): # 加两百个空空格 
        goodsRaw = goodsStartRowNum + i * 2

        # 居中对齐
        for cells in worksheetTemplete.iter_cols(0, 12, goodsRaw, goodsRaw + 1):
            for cell in cells:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置两行的总高度为100
        worksheetTemplete.row_dimensions[goodsRaw].height = 30
        worksheetTemplete.row_dimensions[goodsRaw + 1].height = 30


        if i < len(rawData): # 有数据的表格
            rawDataItem = rawData[i] # 每个订单的原始数据

            # 产品序号
            worksheetTemplete["A{}".format(goodsRaw)] = i + 1

            # 产品名称（写入名称，供 L 列价格公式使用 VLOOKUP 查找）
            worksheetTemplete["C{}".format(goodsRaw)] = rawDataItem["goodsName"]

            # 从 url 获取产品图片
            print("sku：", rawDataItem["sku"], "||color：", rawDataItem["color"], "||图片地址：", rawDataItem["imgUrl"])
            opener = urllib.request.build_opener()
            opener.addheaders = [('User-agent', 'Mozilla/5.0')]
            response = opener.open(rawDataItem["imgUrl"])
            # 暂存图片
            img = PIL.Image.open(io.BytesIO(response.read()))
            img = img.convert("RGB")

            # 没有 temp 文件夹就创建一个 temp 文件夹
            temp_dir = os.path.join(get_base_path(), "temp")
            if not os.path.exists(temp_dir):
                os.mkdir(temp_dir)

            img.save(os.path.join(temp_dir, f"tempImg{i}.jpg"))
            # 设置产品图片
            inster_image(worksheetTemplete, goodsRaw + 1, 2, 0, os.path.join(temp_dir, f"tempImg{i}.jpg"), image_size=(60, 60))

            # 各尺码数量
            sizeNum = rawDataItem["sizeNum"]
            totalNum = 0 # 当前商品的总数
            for eachSize in sizeNum:
                totalNum += sizeNum[eachSize]
                if eachSize == "XS":
                    worksheetTemplete.cell(goodsRaw + 1, goodsTableHead.index("S") + 1, "XS:" + str(sizeNum[eachSize]))
                elif eachSize == "4XL":
                    worksheetTemplete.cell(goodsRaw + 1, goodsTableHead.index("2XL") + 1, "4XL:" + str(sizeNum[eachSize]))
                elif eachSize == "5XL":
                    worksheetTemplete.cell(goodsRaw + 1, goodsTableHead.index("3XL") + 1, "5XL:" + str(sizeNum[eachSize]))
                else:
                    worksheetTemplete.cell(goodsRaw, goodsTableHead.index(eachSize) + 1, str(sizeNum[eachSize]))
            worksheetTemplete.cell(goodsRaw, goodsTableHead.index("合计") + 1, int(totalNum))

        # 复制产品名称单元格的格式
        worksheetTemplete["C{}".format(goodsRaw)].number_format = worksheetTemplete["C9"].number_format
        # worksheetTemplete["C{}".format(goodsRaw)].font = worksheetTemplete["C9"].font.copy()
        # worksheetTemplete["C{}".format(goodsRaw)].alignment = worksheetTemplete["C9"].alignment.copy()
        # worksheetTemplete["C{}".format(goodsRaw)].border = worksheetTemplete["C9"].border.copy()
        # worksheetTemplete["C{}".format(goodsRaw)].fill = worksheetTemplete["C9"].fill.copy()
        # worksheetTemplete["C{}".format(goodsRaw)].protection = worksheetTemplete["C9"].protection.copy()
        worksheetTemplete["C{}".format(goodsRaw)].font = copy.copy(worksheetTemplete["C9"].font)
        worksheetTemplete["C{}".format(goodsRaw)].alignment = copy.copy(worksheetTemplete["C9"].alignment)
        worksheetTemplete["C{}".format(goodsRaw)].border = copy.copy(worksheetTemplete["C9"].border)
        worksheetTemplete["C{}".format(goodsRaw)].fill = copy.copy(worksheetTemplete["C9"].fill)
        worksheetTemplete["C{}".format(goodsRaw)].protection = copy.copy(worksheetTemplete["C9"].protection)

        # 复制数据验证（下拉菜单）
        from openpyxl.worksheet.datavalidation import DataValidation
        # 在工作表的 data_validations 中查找应用于 C9 单元格的验证规则
        source_dv = None
        for dv in worksheetTemplete.data_validations.dataValidation:
            if 'C9' in dv.sqref:
                source_dv = dv
                break
        
        if source_dv:
            new_dv = DataValidation(
                type=source_dv.type,
                operator=source_dv.operator,
                formula1=source_dv.formula1,
                formula2=source_dv.formula2,
                allow_blank=source_dv.allow_blank,
                showDropDown=source_dv.showDropDown,
                showErrorMessage=source_dv.showErrorMessage,
                errorStyle=source_dv.errorStyle,
                errorTitle=source_dv.errorTitle,
                error=source_dv.error,
                showInputMessage=source_dv.showInputMessage,
                promptTitle=source_dv.promptTitle,
                prompt=source_dv.prompt,
            )
            worksheetTemplete.add_data_validation(new_dv)
            new_dv.add(f'C{goodsRaw}:C{goodsRaw + 1}')


        # 合并单元格
        worksheetTemplete.merge_cells(f'B{goodsRaw}:B{goodsRaw + 1}')
        worksheetTemplete.merge_cells('K{}:K{}'.format(goodsRaw, goodsRaw + 1))
        worksheetTemplete.merge_cells('A{}:A{}'.format(goodsRaw, goodsRaw + 1))
        worksheetTemplete.merge_cells(f'C{goodsRaw}:C{goodsRaw + 1}')


        # 填写尺码中的商品数量
        # goodsNum =



        # 设置边框
        from openpyxl.styles.borders import Border, Side
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        for col in range(1, 13):
            for row in range(goodsRaw, goodsRaw + 2):
                cell = worksheetTemplete.cell(row=row, column=col)
                cell.border = thin_border

    # 删除模板中的第9行和第10行数据模板行（在写入数据之后删除）
    # 删除前需要更新公式引用，因为删除行会导致公式引用错位
    # 使用正则表达式更新公式中的单元格引用（删除2行，所以>=11的行号需要减2）
    import re
    for row in range(11, 220):  # 假设最多220行
        for col in range(1, 20):  # 假设最多20列
            cell = worksheetTemplete.cell(row=row, column=col)
            if cell.data_type == 'f' and cell.value:
                # 匹配单元格引用（如J11, C11等），并将>=11的行号减2
                def fix_row(match):
                    col_letter = match.group(1)
                    row_num = int(match.group(2))
                    if row_num >= 11:
                        return f"{col_letter}{row_num - 2}"
                    return match.group(0)
                cell.value = re.sub(r'([A-Za-z])(\d+)', fix_row, cell.value)
    
    # 删除行前调整图片位置（将图片行号减2）
    for img in worksheetTemplete._images:
        anchor = img.anchor
        if hasattr(anchor, '_from') and anchor._from.row >= 10:  # 第10行及以下的图片需要调整
            anchor._from.row -= 2
    
    # 现在删除模板行
    worksheetTemplete.delete_rows(9, 2)

    # 没有 OUTPUT_PATH 这个文件夹则创建这个文件夹
    if not os.path.exists(OUTPUT_PATH):
        os.mkdir(OUTPUT_PATH)

    workbookDianTemplete.save(os.path.join(OUTPUT_PATH, f'{file_name}_result.xlsx'))
    print(os.path.join(OUTPUT_PATH, f'{file_name}_result.xlsx'))


main()

# 暂停程序，等待用户按任意键后退出
try:
    input("按 Enter 键退出...")
except EOFError:
    pass
