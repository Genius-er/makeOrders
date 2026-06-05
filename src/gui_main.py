import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinterdnd2 import TkinterDnD
import os
import sys
import shutil
import tempfile
import threading
import time
import re
import openpyxl
import urllib.request
import io
import PIL.Image
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, XDRPositiveSize2D

class OrderToolGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("做单工具")
        self.root.geometry("634x860")
        self.root.resizable(True, True)

        self.input_file_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()
        self.is_processing = False

        self.create_widgets()

    def create_widgets(self):
        # 创建主容器，带滚动条
        main_canvas = tk.Canvas(self.root)
        main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 创建滚动条
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 配置canvas的滚动
        main_canvas.configure(yscrollcommand=scrollbar.set)
        main_canvas.bind('<Configure>', lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all")))

        # 创建主frame，放在canvas内
        main_frame = ttk.Frame(main_canvas, padding="20")
        main_canvas.create_window((0, 0), window=main_frame, anchor="nw")

        # 当main_frame大小变化时更新canvas的scrollregion
        def on_frame_configure(event):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        main_frame.bind('<Configure>', on_frame_configure)

        # 鼠标滚轮滚动支持
        def on_mouse_wheel(event):
            # 检查鼠标是否在主canvas区域内
            x, y = main_canvas.winfo_pointerxy()
            canvas_x = main_canvas.winfo_rootx()
            canvas_y = main_canvas.winfo_rooty()
            canvas_width = main_canvas.winfo_width()
            canvas_height = main_canvas.winfo_height()
            
            if (canvas_x <= x <= canvas_x + canvas_width and 
                canvas_y <= y <= canvas_y + canvas_height):
                # 检查是否需要滚动（内容高度是否大于窗口高度）
                scroll_region = main_canvas.cget("scrollregion")
                if scroll_region:
                    # scrollregion格式: "0 0 width height"
                    parts = scroll_region.split()
                    if len(parts) == 4:
                        content_height = int(parts[3])
                        # 只有当内容高度大于canvas高度时才允许滚动
                        if content_height <= canvas_height:
                            return 'break'
                
                # 在canvas区域内且需要滚动，处理滚轮事件
                delta = event.delta // 120  # 标准化滚轮增量
                main_canvas.yview_scroll(-delta, "units")
                return 'break'
        
        # 进入canvas区域时绑定滚轮事件
        def on_enter_canvas(event):
            self.root.bind_all('<MouseWheel>', on_mouse_wheel)
        
        # 离开canvas区域时解绑滚轮事件
        def on_leave_canvas(event):
            self.root.unbind_all('<MouseWheel>')
        
        main_canvas.bind('<Enter>', on_enter_canvas)
        main_canvas.bind('<Leave>', on_leave_canvas)

        title_label = ttk.Label(main_frame, text="做单工具", font=("Microsoft YaHei", 20, "bold"))
        title_label.pack(pady=(0, 20))

        instructions_frame = ttk.LabelFrame(main_frame, text="店小秘订单导出说明", padding="10")
        instructions_frame.pack(fill=tk.X, pady=(0, 20))

        instructions_text = """1. 打开店小秘，进入订单页面
2. 勾选要做单的订单
3. 点击"导入/导出"
4. 选择"导出勾选订单"
5. 选择"按产品导出（每个产品导出一行）"
6. 选择"标准模板"
7. 点击"导出"按钮下载文件"""
        instructions_label = ttk.Label(instructions_frame, text=instructions_text, justify=tk.LEFT, font=("Microsoft YaHei", 10))
        instructions_label.pack(fill=tk.X)

        input_frame = ttk.LabelFrame(main_frame, text="选择店小秘导出订单文件", padding="10")
        input_frame.pack(fill=tk.X, pady=(0, 20))

        input_top_frame = ttk.Frame(input_frame)
        input_top_frame.pack(fill=tk.X)

        self.file_entry = ttk.Entry(input_top_frame, textvariable=self.input_file_path, width=50)
        self.file_entry.pack(side=tk.LEFT, padx=(0, 10))

        browse_btn = ttk.Button(input_top_frame, text="浏览...", command=self.browse_file)
        browse_btn.pack(side=tk.LEFT)

        drop_frame = ttk.Frame(input_frame, borderwidth=2, relief=tk.SOLID, padding="20", cursor="hand2")
        drop_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        self.drop_label = ttk.Label(drop_frame, text="将文件拖拽到此处\n或点击上方浏览按钮选择文件", font=("Microsoft YaHei", 12), foreground="gray")
        self.drop_label.pack(expand=True)

        drop_frame.drop_target_register('DND_Files')
        drop_frame.dnd_bind('<<Drop>>', self.on_file_drop)

        output_frame = ttk.LabelFrame(main_frame, text="选择输出保存位置", padding="10")
        output_frame.pack(fill=tk.X, pady=(0, 20))

        output_top_frame = ttk.Frame(output_frame)
        output_top_frame.pack(fill=tk.X)

        self.output_entry = ttk.Entry(output_top_frame, textvariable=self.output_folder_path, width=50)
        self.output_entry.pack(side=tk.LEFT, padx=(0, 10))

        output_btn = ttk.Button(output_top_frame, text="浏览...", command=self.browse_output_folder)
        output_btn.pack(side=tk.LEFT)

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=(0, 20))

        self.start_btn = ttk.Button(button_frame, text="开始执行", command=self.start_processing, width=20)
        self.start_btn.pack(side=tk.LEFT, padx=10)

        self.exit_btn = ttk.Button(button_frame, text="退出", command=self.root.quit, width=20)
        self.exit_btn.pack(side=tk.LEFT, padx=10)

        self.progress_frame = ttk.Frame(main_frame)
        self.progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.progress_label = ttk.Label(self.progress_frame, text="", font=("Microsoft YaHei", 9))
        self.progress_label.pack(anchor=tk.W)

        self.progress_bar = ttk.Progressbar(self.progress_frame, mode='determinate', maximum=100, value=0, length=300)
        self.progress_bar.pack(fill=tk.X, pady=(5, 0))

        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.BOTH, expand=True)

        self.status_text = tk.Text(status_frame, height=8, width=80, state=tk.DISABLED, font=("Consolas", 9))
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        status_scrollbar = ttk.Scrollbar(status_frame, orient="vertical", command=self.status_text.yview)
        status_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.config(yscrollcommand=status_scrollbar.set)

    def on_file_drop(self, event):
        files = self.root.tk.splitlist(event.data)
        for f in files:
            if f.endswith('.xlsx'):
                self.input_file_path.set(f)
                # 设置默认输出文件名：原文件名 + "_工厂单"
                input_basename = os.path.basename(f)
                if input_basename.endswith('.xlsx'):
                    output_filename = input_basename[:-5] + "_工厂单.xlsx"
                else:
                    output_filename = input_basename + "_工厂单.xlsx"
                output_path = os.path.join(os.path.dirname(f), output_filename)
                self.output_folder_path.set(output_path)
                self.update_status(f"已拖入文件: {f}")
                break

    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="选择店小秘导出的订单文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.input_file_path.set(filename)
            # 设置默认输出文件名：原文件名 + "_工厂单"
            input_basename = os.path.basename(filename)
            if input_basename.endswith('.xlsx'):
                output_filename = input_basename[:-5] + "_工厂单.xlsx"
            else:
                output_filename = input_basename + "_工厂单.xlsx"
            output_path = os.path.join(os.path.dirname(filename), output_filename)
            self.output_folder_path.set(output_path)
            self.update_status(f"已选择文件: {filename}")

    def browse_output_folder(self):
        # 获取默认文件名和目录
        default_filename = ""
        initial_dir = "."
        if self.input_file_path.get():
            input_dir = os.path.dirname(self.input_file_path.get())
            if input_dir:
                initial_dir = input_dir
            input_basename = os.path.basename(self.input_file_path.get())
            if input_basename.endswith('.xlsx'):
                default_filename = input_basename[:-5] + "_工厂单.xlsx"
            else:
                default_filename = input_basename + "_工厂单.xlsx"
        
        filename = filedialog.asksaveasfilename(
            title="选择输出文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialdir=initial_dir,
            initialfile=default_filename
        )
        if filename:
            self.output_folder_path.set(filename)
            self.update_status(f"已选择输出文件: {filename}")

    def on_same_folder_toggle(self):
        # 此方法不再需要
        pass

    def update_status(self, message):
        self.status_text.config(state=tk.NORMAL)
        timestamp = time.strftime("%H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.status_text.config(state=tk.DISABLED)

    def set_progress(self, message, progress_value=None):
        self.progress_label.config(text=message)
        if progress_value is not None:
            self.progress_bar.config(mode='determinate', value=progress_value)
            if progress_value == 0:
                self.progress_bar.stop()
        else:
            self.progress_bar.config(mode='indeterminate')
            if not self.is_processing:
                self.progress_bar.stop()
            else:
                self.progress_bar.start(10)

    def validate_inputs(self):
        if not self.input_file_path.get():
            messagebox.showwarning("警告", "请选择店小秘导出的订单文件！")
            return False
        if not os.path.exists(self.input_file_path.get()):
            messagebox.showerror("错误", "选择的文件不存在！")
            return False
        if not self.output_folder_path.get():
            messagebox.showwarning("警告", "请选择输出保存位置！")
            return False
        # 获取输出文件所在的目录
        output_dir = os.path.dirname(self.output_folder_path.get())
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
            except:
                messagebox.showerror("错误", "无法创建输出目录！")
                return False
        return True

    def start_processing(self):
        if self.is_processing:
            return

        if not self.validate_inputs():
            return

        self.is_processing = True
        self.start_btn.config(state=tk.DISABLED)
        self.set_progress("处理中...")

        thread = threading.Thread(target=self.process_orders)
        thread.daemon = True
        thread.start()

    def get_resource_path(self, relative_path):
        try:
            base_path = sys._MEIPASS
        except AttributeError:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def normalize_size(self, size):
        """将非标准码数映射到模板支持的标准码（S, M, L, XL, 2XL, 3XL, 4XL, 5XL）"""
        SIZE_MAPPING = {
            # 西班牙语 / 欧洲码 → 标准码
            "P": "S",
            "CH": "S",
            "G": "L",
            "EG": "XL",
            "EEG": "2XL",
        }
        if size in SIZE_MAPPING:
            mapped = SIZE_MAPPING[size]
            self.update_status(f"码数映射: {size} → {mapped}")
            return mapped
        return size

    def process_orders(self):
        try:
            input_file = self.input_file_path.get()
            output_file = self.output_folder_path.get()

            # 验证输出文件路径
            if not output_file:
                self.root.after(0, lambda: messagebox.showerror("错误", "请选择输出文件"))
                self.root.after(0, lambda: self.start_btn.config(state=tk.NORMAL))
                self.is_processing = False
                return
            
            # 确保输出文件以 .xlsx 结尾
            if not output_file.endswith('.xlsx'):
                output_file = output_file + '.xlsx'
                self.output_folder_path.set(output_file)

            self.update_status("开始处理订单...")
            self.update_status(f"输入文件: {input_file}")
            self.update_status(f"输出文件: {output_file}")

            templete_path = self.get_resource_path(r"templete.xlsx")
            file_name = os.path.basename(input_file)[:-5]

            self.update_status(f"读取模板文件: {templete_path}")
            workbookDianTemplete = openpyxl.load_workbook(templete_path)
            worksheetTemplete = workbookDianTemplete["Sheet1"]

            self.update_status("读取店小秘订单文件...")
            workbookDianXiaoMi = openpyxl.load_workbook(input_file)
            worksheetDianXiaoMi = workbookDianXiaoMi["order_"]

            sheetHead = []
            for cell in worksheetDianXiaoMi[1]:
                sheetHead.append(cell.value)

            self.update_status("解析订单数据...")
            rawData = []
            skuColor = {}

            for row in worksheetDianXiaoMi.iter_rows(min_row=2):
                if all(cell.value is None for cell in row):
                    continue

                productSpecifications = row[sheetHead.index("产品规格")].value

                size = "UNKNOWN"
                color = "UNKNOWN"

                sizeMatches = re.findall(r'Size:(?:xs|s|m|l|xl|xxl|xxxl|xxxxl|2xl|3xl|4xl|5xl)?', productSpecifications, re.IGNORECASE)
                if sizeMatches and sizeMatches[0]:
                    sizePart = sizeMatches[0]
                    size = sizePart.split(":")[1].upper() if len(sizePart.split(":")) > 1 else "UNKNOWN"
                    if size == "":
                        size = "UNKNOWN"
                    colorPart = productSpecifications.replace(sizePart, "")
                    colorMatches = re.findall(r'Color:(.+?)(?:Size|$)', colorPart, re.IGNORECASE)
                    if colorMatches:
                        color = colorMatches[0].strip().replace(" ", "_")
                    else:
                        colorParts = colorPart.split(":")
                        color = colorParts[1].strip().replace(" ", "_") if len(colorParts) > 1 else "UNKNOWN"
                else:
                    sizeMatches = re.findall(r'尺寸[:：]\s*(xs|s|m|l|xl|xxl|xxxl|xxxxl|2xl|3xl|4xl|5xl)', productSpecifications, re.IGNORECASE)
                    if sizeMatches:
                        size = sizeMatches[0].upper()
                    else:
                        sizeMatches = re.findall(r'尺寸[:：]\s*(\w+)', productSpecifications)
                        if sizeMatches:
                            size = sizeMatches[0].upper()

                    colorMatches = re.findall(r'颜色[:：]\s*([^\n\r]+)', productSpecifications)
                    if colorMatches:
                        color = colorMatches[0].strip().replace(" ", "_")

                    # 中文模式未匹配时，尝试英文/西班牙语/意大利语模式 (如 Talla:, Talle:, Color:)
                    if size == "UNKNOWN":
                        # 支持多行格式，如 "Talle:XL\nColor:Gris"
                        sizeMatches = re.findall(r'(?:Talla|Talle|Size|Tallaje)[:：]\s*(xs|s|m|l|xl|xxl|xxxl|xxxxl|2xl|3xl|4xl|5xl)\b', productSpecifications, re.IGNORECASE)
                        if sizeMatches:
                            size = sizeMatches[0].upper()
                        else:
                            # 匹配任意非空白字符作为尺码值（支持多行格式）
                            sizeMatches = re.findall(r'(?:Talla|Talle|Size|Tallaje)[:：]\s*(\S+)', productSpecifications, re.IGNORECASE)
                            if sizeMatches:
                                size = sizeMatches[0].upper()

                    if color == "UNKNOWN":
                        # 支持多行格式，如 "Talle:XL\nColor:Gris"
                        colorMatches = re.findall(r'Color[:：]\s*([^\n\r]+)', productSpecifications, re.IGNORECASE)
                        if colorMatches:
                            color = colorMatches[0].strip().replace(" ", "_")
                        else:
                            # 尝试匹配 Color: 后面直到换行或行尾的内容
                            colorMatches = re.findall(r'Color[:：]\s*(\S+)', productSpecifications, re.IGNORECASE)
                            if colorMatches:
                                color = colorMatches[0].strip().replace(" ", "_")

                num_x = len(re.findall(r'X', size))
                if num_x > 1 and size != "UNKNOWN":
                    size = size.replace('X' * num_x, str(num_x) + "X")

                # 将非标准码数映射到模板支持的标准码
                size = self.normalize_size(size)

                if size == "UNKNOWN" and color == "UNKNOWN":
                    self.update_status(f"警告：无法解析产品规格 '{productSpecifications}'，跳过此行")
                    continue

                goodsId = row[sheetHead.index("产品ID")].value
                rawDataItem = {}

                goodsNum = row[sheetHead.index("产品总数")].value
                if goodsId + color in skuColor:
                    rawDataItem = rawData[skuColor[goodsId+color]]
                    if size in rawDataItem["sizeNum"]:
                        rawDataItem["sizeNum"][size] = int(rawDataItem["sizeNum"][size] + goodsNum)
                    else:
                        rawDataItem["sizeNum"][size] = int(goodsNum)
                else:
                    rawDataItem = {
                        "goodsName": row[sheetHead.index("产品名称")].value,
                        "sizeNum": {size: int(goodsNum)},
                        "imgUrl": row[sheetHead.index("图片网址")].value,
                        "sku": goodsId,
                        "color": color,
                    }
                    skuColor[goodsId + color] = len(rawData)
                    rawData.append(rawDataItem)

            self.update_status(f"共解析 {len(rawData)} 个商品")

            goodsStartRowNum = 9

            goodsTableHead = []
            for cell in worksheetTemplete[7]:
                goodsTableHead.append(cell.value)

            for i in range(len(rawData) + 200):
                goodsRaw = goodsStartRowNum + i * 2

                for cells in worksheetTemplete.iter_cols(0, 12, goodsRaw, goodsRaw + 1):
                    for cell in cells:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                worksheetTemplete.row_dimensions[goodsRaw].height = 30
                worksheetTemplete.row_dimensions[goodsRaw + 1].height = 30

                if i < len(rawData):
                    rawDataItem = rawData[i]

                    worksheetTemplete["A{}".format(goodsRaw)] = i + 1

                    self.update_status(f"下载图片: {rawDataItem['sku']} - {rawDataItem['color']}")
                    opener = urllib.request.build_opener()
                    opener.addheaders = [('User-agent', 'Mozilla/5.0')]
                    response = opener.open(rawDataItem["imgUrl"])
                    img = PIL.Image.open(io.BytesIO(response.read()))
                    img = img.convert("RGB")

                    temp_dir = os.path.join(tempfile.gettempdir(), "order_tool_temp")
                    if not os.path.exists(temp_dir):
                        os.makedirs(temp_dir)
                    img_path = os.path.join(temp_dir, f"tempImg{i}.jpg")
                    img.save(img_path)

                    self.insert_image(worksheetTemplete, goodsRaw + 1, 2, 0, img_path, image_size=(60, 60))

                    sizeNum = rawDataItem["sizeNum"]
                    totalNum = 0
                    for eachSize in sizeNum:
                        totalNum += sizeNum[eachSize]
                        if eachSize == "XS":
                            worksheetTemplete.cell(goodsRaw + 1, goodsTableHead.index("S") + 1, "XS:" + str(sizeNum[eachSize]))
                        elif eachSize == "4XL":
                            worksheetTemplete.cell(goodsRaw + 1, goodsTableHead.index("2XL") + 1, "4XL:" + str(sizeNum[eachSize]))
                        elif eachSize == "5XL":
                            worksheetTemplete.cell(goodsRaw + 1, goodsTableHead.index("3XL") + 1, "5XL:" + str(sizeNum[eachSize]))
                        else:
                            worksheetTemplete.cell(goodsRaw, goodsTableHead.index(eachSize) + 1, sizeNum[eachSize])
                    # 合计列使用简单 SUM 公式，D-I 列已确保填入数字类型
                    # 如果求和为 0 则显示为空，否则显示合计值
                    worksheetTemplete.cell(goodsRaw, goodsTableHead.index("合计") + 1, f'=IF(SUM(D{goodsRaw}:I{goodsRaw})=0,"",SUM(D{goodsRaw}:I{goodsRaw}))')
                    worksheetTemplete.cell(goodsRaw + 1, goodsTableHead.index("合计") + 1, f'=IF(SUM(D{goodsRaw + 1}:I{goodsRaw + 1})=0,"",SUM(D{goodsRaw + 1}:I{goodsRaw + 1}))')

                    # 复制 L 列的价格公式（从模板行复制）
                    # 第一行（goodsRaw）
                    template_formula_row1 = worksheetTemplete.cell(row=9, column=12).value  # L9 的公式
                    if template_formula_row1 and template_formula_row1.startswith('='):
                        new_formula_row1 = template_formula_row1.replace('J9', f'J{goodsRaw}').replace('C9', f'C{goodsRaw}')
                        worksheetTemplete.cell(row=goodsRaw, column=12).value = new_formula_row1
                    
                    # 第二行（goodsRaw + 1）- J列引用第二行（用户填写数量），C列引用第一行（因为C列是合并的）
                    template_formula_row2 = worksheetTemplete.cell(row=10, column=12).value  # L10 的公式
                    if template_formula_row2 and template_formula_row2.startswith('='):
                        new_formula_row2 = template_formula_row2.replace('J10', f'J{goodsRaw + 1}').replace('C10', f'C{goodsRaw}')
                        worksheetTemplete.cell(row=goodsRaw + 1, column=12).value = new_formula_row2

                worksheetTemplete["C{}".format(goodsRaw)].number_format = worksheetTemplete["C9"].number_format
                worksheetTemplete["C{}".format(goodsRaw)].font = worksheetTemplete["C9"].font.copy()
                worksheetTemplete["C{}".format(goodsRaw)].alignment = worksheetTemplete["C9"].alignment.copy()
                worksheetTemplete["C{}".format(goodsRaw)].border = worksheetTemplete["C9"].border.copy()
                worksheetTemplete["C{}".format(goodsRaw)].fill = worksheetTemplete["C9"].fill.copy()

                from openpyxl.worksheet.datavalidation import DataValidation
                source_dv = None
                for dv in worksheetTemplete.data_validations.dataValidation:
                    if 'C9' in dv.sqref:
                        source_dv = dv
                        break

                if source_dv:
                    new_dv = DataValidation(
                        type=source_dv.type,
                        operator=source_dv.operator,
                        formula1=source_dv.formula1,
                        formula2=source_dv.formula2,
                        allow_blank=source_dv.allow_blank,
                        showDropDown=source_dv.showDropDown,
                        showErrorMessage=source_dv.showErrorMessage,
                        errorStyle=source_dv.errorStyle,
                        errorTitle=source_dv.errorTitle,
                        error=source_dv.error,
                        showInputMessage=source_dv.showInputMessage,
                        promptTitle=source_dv.promptTitle,
                        prompt=source_dv.prompt,
                    )
                    worksheetTemplete.add_data_validation(new_dv)
                    new_dv.add(f'C{goodsRaw}:C{goodsRaw + 1}')

                worksheetTemplete.merge_cells(f'B{goodsRaw}:B{goodsRaw + 1}')
                worksheetTemplete.merge_cells('K{}:K{}'.format(goodsRaw, goodsRaw + 1))
                worksheetTemplete.merge_cells('A{}:A{}'.format(goodsRaw, goodsRaw + 1))
                worksheetTemplete.merge_cells(f'C{goodsRaw}:C{goodsRaw + 1}')

                thin_border = Border(left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))
                for col in range(1, 13):
                    for row_idx in range(goodsRaw, goodsRaw + 2):
                        cell = worksheetTemplete.cell(row=row_idx, column=col)
                        cell.border = thin_border

            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)

            # 如果目标文件已存在且是文件夹，先删除它
            if os.path.exists(output_file) and os.path.isdir(output_file):
                shutil.rmtree(output_file)
            # 如果目标文件已存在且是文件，先删除它
            elif os.path.exists(output_file) and os.path.isfile(output_file):
                os.remove(output_file)

            workbookDianTemplete.save(output_file)

            self.update_status(f"结果已保存到: {output_file}")
            self.root.after(0, lambda: messagebox.showinfo("完成", f"订单处理完成！\n结果已保存到:\n{output_file}"))

            shutil.rmtree(temp_dir, ignore_errors=True)

        except Exception as e:
            self.update_status(f"错误: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("错误", f"处理订单时出错:\n{str(e)}"))

        finally:
            self.is_processing = False
            self.root.after(0, self.set_progress, "", 0)
            self.root.after(0, lambda: self.start_btn.config(state=tk.NORMAL))

    def insert_image(self, worksheet, start_row, start_col, hieght, image_url, image_size=None):
        img = XLImage(image_url)
        img.height, img.width = image_size
        col_letter = get_column_letter(start_col)
        width = worksheet.column_dimensions[col_letter].width
        c2e = cm_to_EMU
        p2e = pixels_to_EMU
        size = XDRPositiveSize2D(p2e(img.height), p2e(img.width))
        image_cell, mod = divmod(img.height, 20)
        image_cell = image_cell + 1 if mod != 0 else image_cell
        if (hieght - image_cell) % 2 == 0:
            start = start_row + (hieght - image_cell) // 2
            rowOff = c2e((0.6 * 49.77) / 99)
        else:
            start = start_row + (hieght - image_cell) // 2 + 1
            rowOff = c2e((0.4 * 49.77) / 99)
        colOff = (c2e(width + 1) - c2e(img.width / 72 * 13)) / 9
        marker = AnchorMarker(col=start_col - 1, colOff=colOff, row=start - 1, rowOff=rowOff)
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        worksheet.add_image(img)

def main():
    root = TkinterDnD.Tk()
    app = OrderToolGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
